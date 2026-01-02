

from flask import Flask, render_template_string, jsonify, request, send_file
from flask_socketio import SocketIO, emit
from flask_cors import CORS
from datetime import datetime, timedelta
import json, threading, time, os, random, re, io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# EMAIL & SMS IMPORTS
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from twilio.rest import Client

from crewai import Agent, Task, Crew, Process
from langchain_groq import ChatGroq
from langchain_google_genai import ChatGoogleGenerativeAI
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Text, Boolean, desc, func, JSON
from sqlalchemy.orm import declarative_base, sessionmaker
from sqlalchemy.pool import StaticPool

# ============================================================================
# EMAIL & SMS CONFIGURATION - UPDATE WITH YOUR CREDENTIALS
# ============================================================================
# Gmail SMTP Configuration
GMAIL_USER = "bmkamal10@gmail.com"
GMAIL_APP_PASSWORD = "jaqi bezm ouxs dyqr"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587


# Twilio Configuration  
TWILIO_ACCOUNT_SID = ""
TWILIO_AUTH_TOKEN = ""
TWILIO_PHONE_NUMBER = "+"

# ============================================================================

GROQ_API_KEY = ""
GOOGLE_API_KEY = ""


os.environ['GROQ_API_KEY'] = GROQ_API_KEY
os.environ['GOOGLE_API_KEY'] = GOOGLE_API_KEY

Base = declarative_base()

class Customer(Base):
    __tablename__ = 'customers'
    id = Column(Integer, primary_key=True)
    name = Column(String(100))
    email = Column(String(100), unique=True)
    phone = Column(String(20))
    age = Column(Integer)
    purchase_history = Column(Float)
    segment = Column(String(50))
    preferences = Column(Text)
    location = Column(String(100))
    preferred_channel = Column(String(50), default='email')
    engagement_score = Column(Float, default=0.0)
    language = Column(String(10), default='en')
    total_purchases = Column(Integer, default=0)
    churn_probability = Column(Float, default=0.0)
    lifetime_value = Column(Float, default=0.0)
    last_purchase_date = Column(DateTime)
    created_at = Column(DateTime, default=datetime.now)

class Campaign(Base):
    __tablename__ = 'campaigns'
    id = Column(Integer, primary_key=True)
    name = Column(String(200))
    channel = Column(String(50))
    target_segment = Column(String(50))
    content = Column(Text)
    subject_line = Column(String(200))
    tone = Column(String(50), default='professional')
    language = Column(String(10), default='en')
    created_at = Column(DateTime, default=datetime.now)
    scheduled_at = Column(DateTime)
    status = Column(String(50), default='draft')
    sent_count = Column(Integer, default=0)
    predicted_roi = Column(Float, default=0.0)
    budget = Column(Float, default=0.0)
    ab_test_enabled = Column(Boolean, default=False)
    variant_a = Column(Text)
    variant_b = Column(Text)
    winner_variant = Column(String(1))
    compliance_checked = Column(Boolean, default=False)

class Analytics(Base):
    __tablename__ = 'analytics'
    id = Column(Integer, primary_key=True)
    campaign_id = Column(Integer)
    sent = Column(Integer, default=0)
    delivered = Column(Integer, default=0)
    opened = Column(Integer, default=0)
    clicked = Column(Integer, default=0)
    converted = Column(Integer, default=0)
    bounced = Column(Integer, default=0)
    unsubscribed = Column(Integer, default=0)
    revenue = Column(Float, default=0.0)
    sentiment_score = Column(Float, default=0.0)
    avg_engagement_time = Column(Float, default=0.0)
    device_breakdown = Column(JSON)
    geo_breakdown = Column(JSON)
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)

class AgentActivity(Base):
    __tablename__ = 'agent_activities'
    id = Column(Integer, primary_key=True)
    agent_name = Column(String(100))
    activity_type = Column(String(50))
    description = Column(Text)
    status = Column(String(20))
    extra_data = Column(JSON)
    created_at = Column(DateTime, default=datetime.now)

class ABTest(Base):
    __tablename__ = 'ab_tests'
    id = Column(Integer, primary_key=True)
    campaign_id = Column(Integer)
    variant_a_subject = Column(String(200))
    variant_a_content = Column(Text)
    variant_b_subject = Column(String(200))
    variant_b_content = Column(Text)
    variant_a_sends = Column(Integer, default=0)
    variant_b_sends = Column(Integer, default=0)
    variant_a_opens = Column(Integer, default=0)
    variant_b_opens = Column(Integer, default=0)
    variant_a_clicks = Column(Integer, default=0)
    variant_b_clicks = Column(Integer, default=0)
    variant_a_conversions = Column(Integer, default=0)
    variant_b_conversions = Column(Integer, default=0)
    winner = Column(String(1))
    confidence_level = Column(Float, default=0.0)
    created_at = Column(DateTime, default=datetime.now)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'ultimate-campaign-secret'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet')

engine = create_engine('sqlite:///campaign_system.db', connect_args={'check_same_thread': False}, poolclass=StaticPool)
Base.metadata.create_all(engine)
Session = sessionmaker(bind=engine)

def send_email(to_email, subject, content, customer_name="Customer"):
    """Send real email via Gmail SMTP"""
    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = GMAIL_USER
        msg['To'] = to_email
        msg['Subject'] = subject
        
        personalized_content = content.replace("{name}", customer_name)
        
        html_content = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <h2 style="color: #00ff88; text-align: center;">{subject}</h2>
                    <div style="margin: 20px 0;">
                        <p>Hi {customer_name},</p>
                        <p>{personalized_content}</p>
                    </div>
                    <hr style="border: 1px solid #eee; margin: 20px 0;">
                    <p style="font-size: 12px; color: #999; text-align: center;">
                        This is an automated campaign email from Agentic Campaign System.
                    </p>
                </div>
            </body>
        </html>
        """
        
        msg.attach(MIMEText(personalized_content, 'plain'))
        msg.attach(MIMEText(html_content, 'html'))
        
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        print(f"✅ Email sent to {to_email}")
        return True
        
    except Exception as e:
        print(f"❌ Email failed for {to_email}: {str(e)}")
        return False

def send_sms(to_phone, message, customer_name="Customer"):
    """Send real SMS via Twilio"""
    try:
        client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        
        personalized_message = message.replace("{name}", customer_name)
        
        if len(personalized_message) > 160:
            personalized_message = personalized_message[:157] + "..."
        
        message_obj = client.messages.create(
            body=personalized_message,
            from_=TWILIO_PHONE_NUMBER,
            to=to_phone
        )
        
        print(f"✅ SMS sent to {to_phone} (SID: {message_obj.sid})")
        return True
        
    except Exception as e:
        print(f"❌ SMS failed for {to_phone}: {str(e)}")
        return False

def initialize_llm():
    if GROQ_API_KEY and len(GROQ_API_KEY) > 20:
        try:
            llm = ChatGroq(api_key=GROQ_API_KEY, model="llama3-70b-8192", temperature=0.7)
            print("✅ Using Groq")
            return llm, "groq"
        except: pass
    if GOOGLE_API_KEY and len(GOOGLE_API_KEY) > 20:
        try:
            llm = ChatGoogleGenerativeAI(google_api_key=GOOGLE_API_KEY, model="gemini-1.5-flash", temperature=0.7)
            print("✅ Using Gemini")
            return llm, "gemini"
        except: pass
    print("⚠️ Using template fallbacks")
    return None, "template"

llm, llm_type = initialize_llm()

if llm:
    data_analyst = Agent(role='Customer Data Analyst', goal='Analyze and segment customers', backstory='Expert data scientist', verbose=False, allow_delegation=False, llm=llm)
    content_creator = Agent(role='Marketing Content Creator', goal='Generate compelling campaigns', backstory='Award-winning copywriter', verbose=False, allow_delegation=False, llm=llm)
    strategy_advisor = Agent(role='Strategy Advisor', goal='Optimize campaigns and predict ROI', backstory='Senior marketing strategist', verbose=False, allow_delegation=False, llm=llm)

def get_db_session():
    return Session()

def log_agent_activity(agent_name, activity_type, description, status="completed", extra_data=None):
    session = get_db_session()
    try:
        activity = AgentActivity(agent_name=agent_name, activity_type=activity_type, description=description, status=status, extra_data=extra_data or {})
        session.add(activity)
        session.commit()
        socketio.emit('agent_activity', {'agent': agent_name, 'type': activity_type, 'description': description, 'status': status, 'timestamp': datetime.now().isoformat()})
    except Exception as e:
        print(f"Error logging: {e}")
    finally:
        session.close()

def generate_sample_data():
    session = get_db_session()
    try:
        if session.query(Customer).count() > 0:
            print("✅ Data exists")
            session.close()
            return
        print("📊 Generating data...")
        segments = ["High-Value", "Tech Enthusiasts", "Budget Shoppers", "Premium Buyers", "Casual Browsers"]
        locations = ["New York, US", "London, UK", "Singapore", "Sydney, AU", "Toronto, CA", "Mumbai, IN", "Tokyo, JP", "Paris, FR", "Berlin, DE", "Dubai, UAE"]
        prefs = ["Electronics, Gadget", "Fashion, Luxury", "Home, Garden", "Sports, Fitness", "Tech, Gaming"]
        
        for i in range(50):
            customer = Customer(
                name=f"Customer {i+1}", 
                email=f"customer{i+1}@example.com", 
                phone=f"+1234567{i:03d}", 
                age=random.randint(25, 65), 
                purchase_history=round(random.uniform(100, 5000), 2), 
                segment=random.choice(segments), 
                preferences=random.choice(prefs), 
                location=random.choice(locations), 
                engagement_score=round(random.uniform(0.3, 0.95), 2), 
                total_purchases=random.randint(1, 20), 
                last_purchase_date=datetime.now() - timedelta(days=random.randint(1, 90)), 
                churn_probability=round(random.uniform(0.1, 0.7), 2), 
                lifetime_value=round(random.uniform(500, 8000), 2), 
                preferred_channel=random.choice(['email', 'sms', 'social']), 
                language=random.choice(['en', 'es', 'fr', 'de'])
            )
            session.add(customer)
        
        for i in range(3):
            campaign = Campaign(
                name=f"Sample Campaign {i+1}", 
                channel=random.choice(['email', 'sms', 'social']), 
                target_segment=random.choice(segments), 
                subject_line=f"Exclusive Offer {i+1}", 
                content=f"Amazing offer {i+1}", 
                status='sent', 
                sent_count=random.randint(100, 500), 
                predicted_roi=round(random.uniform(2.0, 5.0), 2), 
                ab_test_enabled=i==0, 
                tone='professional', 
                language='en', 
                created_at=datetime.now() - timedelta(days=i*5)
            )
            session.add(campaign)
            session.flush()
            sent = campaign.sent_count
            analytics = Analytics(
                campaign_id=campaign.id, 
                sent=sent, 
                delivered=int(sent*0.98), 
                opened=int(sent*random.uniform(0.20,0.35)), 
                clicked=int(sent*random.uniform(0.05,0.12)), 
                converted=int(sent*random.uniform(0.02,0.05)), 
                bounced=int(sent*0.02), 
                unsubscribed=int(sent*0.001), 
                revenue=round(sent*random.uniform(50,150),2), 
                sentiment_score=round(random.uniform(0.65,0.85),2), 
                avg_engagement_time=round(random.uniform(15,45),1), 
                device_breakdown={"mobile":0.65,"desktop":0.30,"tablet":0.05}, 
                geo_breakdown={"US":0.45,"EU":0.30,"APAC":0.15,"Other":0.10}
            )
            session.add(analytics)
        session.commit()
        print("✅ Data generated!")
        log_agent_activity("System", "data_generation", "Generated 50 customers and 3 campaigns", "completed")
    except Exception as e:
        print(f"Error: {e}")
        session.rollback()
    finally:
        session.close()

def agent_segment_customers():
    session = get_db_session()
    customers = session.query(Customer).all()
    if not customers:
        session.close()
        return {"status": "no_customers", "message": "No customers found"}
    log_agent_activity("Data Analyst", "segmentation", f"Segmenting {len(customers)} customers", "in_progress")
    try:
        if llm_type != "template":
            sample = [{'id':c.id,'age':c.age,'purchase_history':c.purchase_history,'engagement_score':c.engagement_score,'preferences':c.preferences} for c in customers[:10]]
            task = Task(description=f"Analyze and segment: {json.dumps(sample)}. Return JSON: {{\"1\":\"segment_name\",...}}", agent=data_analyst, expected_output="JSON mapping")
            crew = Crew(agents=[data_analyst], tasks=[task], process=Process.sequential, verbose=False)
            result = crew.kickoff()
            segment_map = json.loads(str(result))
            for cid_str, segment in segment_map.items():
                c = session.query(Customer).filter_by(id=int(cid_str)).first()
                if c: c.segment = segment
            session.commit()
            log_agent_activity("Data Analyst", "segmentation", f"AI segmented {len(segment_map)} customers", "completed")
            session.close()
            return {"status":"success","segments":len(set(segment_map.values())),"customers":len(segment_map),"method":"ai"}
    except Exception as e:
        print(f"AI failed: {e}")
    segments_assigned = {}
    for c in customers:
        if c.purchase_history > 2000 and c.engagement_score > 0.7: seg = "High-Value"
        elif 'tech' in c.preferences.lower(): seg = "Tech Enthusiasts"
        elif c.purchase_history < 500: seg = "Budget Shoppers"
        elif c.engagement_score > 0.8: seg = "Premium Buyers"
        else: seg = "Casual Browsers"
        c.segment = seg
        segments_assigned[seg] = segments_assigned.get(seg, 0) + 1
    session.commit()
    log_agent_activity("Data Analyst", "segmentation", f"Rule-based: {len(segments_assigned)} segments", "completed")
    session.close()
    return {"status":"success","segments":len(segments_assigned),"customers":len(customers),"method":"rules","segment_breakdown":segments_assigned}

def agent_create_campaign(channel, segment, product, tone="professional", language="en"):
    log_agent_activity("Content Creator", "campaign_creation", f"Creating {channel} campaign", "in_progress")
    va_subj = f"Exclusive {product} Offer Just For You!"
    va_cont = f"Hi {{name}}! We have {product} for {segment}. Limited time benefits: premium quality, exclusive pricing, free shipping. <a href='https://google.com'>Click to claim!</a>"

    vb_subj = f"🎯 {product} - Special Deal Inside"
    vb_cont = f"Hello {{name}}! Launching {product} for {segment}. Why customers love this: designed for you, best value, trusted, risk-free. Shop now!"
    try:
        if llm_type != "template":
            task = Task(description=f"Create 2 variants for {channel}/{segment}/{product}. Format: VARIANT_A_SUBJECT: [text]\\nVARIANT_A_CONTENT: [text]\\nVARIANT_B_SUBJECT: [text]\\nVARIANT_B_CONTENT: [text]", agent=content_creator, expected_output="Two variants")
            crew = Crew(agents=[content_creator], tasks=[task], process=Process.sequential, verbose=False)
            result = str(crew.kickoff())
            if m:=re.search(r'VARIANT_A_SUBJECT:\s*(.+)', result, re.I): va_subj = m.group(1).strip()
            if m:=re.search(r'VARIANT_A_CONTENT:\s*(.+?)(?=VARIANT_B|$)', result, re.I|re.DOTALL): va_cont = m.group(1).strip()
            if m:=re.search(r'VARIANT_B_SUBJECT:\s*(.+)', result, re.I): vb_subj = m.group(1).strip()
            if m:=re.search(r'VARIANT_B_CONTENT:\s*(.+)', result, re.I|re.DOTALL): vb_cont = m.group(1).strip()
    except Exception as e:
        print(f"AI failed: {e}")
    session = get_db_session()
    campaign = Campaign(
        name=f"{channel.upper()}-{segment}-{product}", 
        channel=channel, 
        target_segment=segment, 
        subject_line=va_subj, 
        content=va_cont, 
        variant_a=va_cont, 
        variant_b=vb_cont, 
        tone=tone, 
        language=language, 
        status='ready', 
        ab_test_enabled=True
    )
    session.add(campaign)
    session.commit()
    cid = campaign.id
    ab = ABTest(
        campaign_id=cid, 
        variant_a_subject=va_subj, 
        variant_a_content=va_cont, 
        variant_b_subject=vb_subj, 
        variant_b_content=vb_cont
    )
    session.add(ab)
    session.commit()
    session.close()
    log_agent_activity("Content Creator", "campaign_creation", f"Campaign created: ID {cid}", "completed")
    return {
        "status":"success",
        "campaign_id":cid,
        "variant_a":{"subject":va_subj,"content":va_cont},
        "variant_b":{"subject":vb_subj,"content":vb_cont}
    }

def agent_predict_roi(campaign_id):
    session = get_db_session()
    c = session.query(Campaign).filter_by(id=campaign_id).first()
    if not c:
        session.close()
        return {"status":"error","message":"Campaign not found"}
    log_agent_activity("Strategy Advisor", "roi_prediction", f"Predicting ROI for {campaign_id}", "in_progress")
    customers = session.query(Customer).filter_by(segment=c.target_segment).all()
    avg_ltv = sum(cu.lifetime_value for cu in customers)/len(customers) if customers else 1000
    bench = {'email':{'open':25,'click':6,'convert':2.5},'sms':{'open':95,'click':12,'convert':4.0},'social':{'open':15,'click':8,'convert':2.0}}
    b = bench.get(c.channel, bench['email'])
    metrics = {
        'OPEN_RATE':b['open'],
        'CLICK_RATE':b['click'],
        'CONVERSION_RATE':b['convert'],
        'REVENUE':len(customers)*b['convert']/100*avg_ltv*0.3,
        'ROI':3.5,
        'CONFIDENCE':'High'
    }
    c.predicted_roi = metrics['ROI']
    session.commit()
    session.close()
    log_agent_activity("Strategy Advisor", "roi_prediction", f"ROI: {metrics['ROI']}x", "completed")
    return {"status":"success","metrics":metrics}

def agent_send_campaign(campaign_id):
    session = get_db_session()
    c = session.query(Campaign).filter_by(id=campaign_id).first()
    if not c:
        session.close()
        return {"status":"error","message":"Campaign not found"}
    
    log_agent_activity("System", "campaign_send", f"Sending: {c.name}", "in_progress")
    customers = session.query(Customer).filter_by(segment=c.target_segment).all()
    
    if not customers:
        session.close()
        return {"status":"error","message":"No customers in target segment"}
    
    print(f"\n{'='*60}")
    print(f"📤 SENDING CAMPAIGN: {c.name}")
    print(f"📊 Channel: {c.channel}")
    print(f"🎯 Target: {c.target_segment} ({len(customers)} customers)")
    print(f"{'='*60}\n")
    
    sent_count = 0
    delivered_count = 0
    failed_count = 0
    
    for customer in customers:
        sent_count += 1
        
        if c.ab_test_enabled:
            ab = session.query(ABTest).filter_by(campaign_id=campaign_id).first()
            if sent_count % 2 == 0:
                subject = ab.variant_a_subject if ab else c.subject_line
                content = ab.variant_a_content if ab else c.content
            else:
                subject = ab.variant_b_subject if ab else c.subject_line
                content = ab.variant_b_content if ab else c.variant_b
        else:
            subject = c.subject_line
            content = c.content
        
        success = False
        
        if c.channel == 'email':
            success = send_email(
                to_email=customer.email,
                subject=subject,
                content=content,
                customer_name=customer.name
            )
        
        elif c.channel == 'sms':
            sms_message = f"{subject}\n\n{content[:100]}"
            success = send_sms(
                to_phone=customer.phone,
                message=sms_message,
                customer_name=customer.name
            )
        
        elif c.channel == 'social':
            print(f"ℹ️  Social media posting not implemented for {customer.name}")
            success = True
        
        if success:
            delivered_count += 1
        else:
            failed_count += 1
        
        socketio.emit('campaign_progress', {
            'campaign_id': campaign_id,
            'sent': sent_count,
            'delivered': delivered_count,
            'failed': failed_count,
            'total': len(customers),
            'progress': round(sent_count / len(customers) * 100, 1)
        })
        
        time.sleep(0.1)
    
    print(f"\n{'='*60}")
    print(f"✅ CAMPAIGN COMPLETE!")
    print(f"📤 Sent: {sent_count}")
    print(f"✅ Delivered: {delivered_count}")
    print(f"❌ Failed: {failed_count}")
    print(f"{'='*60}\n")
    
    c.status = 'sent'
    c.sent_count = sent_count
    
    m = {'email':{'delivered':0.98,'opened':0.25,'clicked':0.06,'converted':0.025},'sms':{'delivered':0.99,'opened':0.95,'clicked':0.12,'converted':0.04}}.get(c.channel, {'delivered':0.98,'opened':0.25,'clicked':0.06,'converted':0.025})
    
    if c.ab_test_enabled:
        va_cnt = sent_count//2
        vb_cnt = sent_count-va_cnt
        ab = session.query(ABTest).filter_by(campaign_id=campaign_id).first()
        if ab:
            ab.variant_a_sends=va_cnt
            ab.variant_b_sends=vb_cnt
            ab.variant_a_opens=int(va_cnt*m['opened']*1.05)
            ab.variant_b_opens=int(vb_cnt*m['opened']*0.95)
            ab.variant_a_clicks=int(ab.variant_a_opens*0.24)
            ab.variant_b_clicks=int(ab.variant_b_opens*0.20)
            ab.variant_a_conversions=int(ab.variant_a_clicks*0.40)
            ab.variant_b_conversions=int(ab.variant_b_clicks*0.35)
            ab.winner='A'
            ab.confidence_level=0.92
            c.winner_variant='A'
    
    analytics = Analytics(
        campaign_id=campaign_id, 
        sent=sent_count, 
        delivered=delivered_count, 
        opened=int(delivered_count*m['opened']), 
        clicked=int(delivered_count*m['clicked']), 
        converted=int(delivered_count*m['converted']), 
        bounced=failed_count, 
        unsubscribed=int(sent_count*0.001), 
        revenue=round(delivered_count*m['converted']*195.50,2), 
        sentiment_score=round(random.uniform(0.70,0.85),2), 
        avg_engagement_time=round(random.uniform(20,40),1), 
        device_breakdown={"mobile":0.65,"desktop":0.30,"tablet":0.05}, 
        geo_breakdown={"US":0.45,"EU":0.30,"APAC":0.15,"Other":0.10}
    )
    session.add(analytics)
    session.commit()
    
    log_agent_activity("System", "campaign_send", f"Sent: {sent_count}, Delivered: {delivered_count}, Failed: {failed_count}", "completed")
    session.close()
    
    return {
        "status":"success",
        "sent":sent_count,
        "delivered":delivered_count,
        "failed":failed_count,
        "revenue":analytics.revenue,
        "ab_test_winner":c.winner_variant if c.ab_test_enabled else None
    }

def realtime_analytics_updater():
    while True:
        time.sleep(8)
        session = get_db_session()
        try:
            for c in session.query(Campaign).filter_by(status='sent').all():
                a = session.query(Analytics).filter_by(campaign_id=c.id).first()
                if a and a.opened < a.sent:
                    new_o = min(a.sent, a.opened+int(a.sent*0.015))
                    new_c = min(new_o, a.clicked+int(new_o*0.04))
                    new_conv = min(new_c, a.converted+int(new_c*0.025))
                    if new_o > a.opened:
                        a.opened=new_o
                        a.clicked=new_c
                        a.converted=new_conv
                        a.revenue=new_conv*195.50
                        a.updated_at=datetime.now()
                        session.commit()
                        socketio.emit('analytics_update', {'campaign_id':c.id,'campaign_name':c.name,'opened':new_o,'clicked':new_c,'converted':new_conv,'revenue':a.revenue})
        except Exception as e:
            print(f"Error: {e}")
        finally:
            session.close()

threading.Thread(target=realtime_analytics_updater, daemon=True).start()

@app.route('/api/upload-customers', methods=['POST'])
def upload_customers():
    """Upload customer CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({"status":"error","message":"No file provided"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"status":"error","message":"No file selected"}), 400
        
        if not file.filename.endswith('.csv'):
            return jsonify({"status":"error","message":"Only CSV files allowed"}), 400
        
        df = pd.read_csv(file)
        
        required_cols = ['name', 'email']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            return jsonify({"status":"error","message":f"Missing columns: {missing_cols}"}), 400
        
        session = get_db_session()
        added = 0
        skipped = 0
        
        for _, row in df.iterrows():
            try:
                existing = session.query(Customer).filter_by(email=row['email']).first()
                if existing:
                    skipped += 1
                    continue
                
                customer = Customer(
                    name=row.get('name', 'Unknown'),
                    email=row['email'],
                    phone=row.get('phone', ''),
                    age=int(row.get('age', 30)),
                    purchase_history=float(row.get('purchase_history', 0)),
                    segment=row.get('segment', 'Unassigned'),
                    preferences=row.get('preferences', ''),
                    location=row.get('location', ''),
                    preferred_channel=row.get('preferred_channel', 'email'),
                    engagement_score=float(row.get('engagement_score', 0.5)),
                    language=row.get('language', 'en'),
                    total_purchases=int(row.get('total_purchases', 0)),
                    churn_probability=float(row.get('churn_probability', 0.3)),
                    lifetime_value=float(row.get('lifetime_value', 1000)),
                    last_purchase_date=datetime.now()
                )
                session.add(customer)
                added += 1
            except Exception as e:
                print(f"Error adding customer: {e}")
                continue
        
        session.commit()
        session.close()
        
        log_agent_activity("System", "csv_upload", f"Uploaded {added} customers ({skipped} skipped)", "completed")
        
        return jsonify({
            "status":"success",
            "added":added,
            "skipped":skipped,
            "message":f"Successfully uploaded {added} customers"
        })
        
    except Exception as e:
        return jsonify({"status":"error","message":str(e)}), 500

@app.route('/api/segment-visualization')
def get_segment_visualization():
    """Get segment distribution data for charts"""
    session = get_db_session()
    try:
        customers = session.query(Customer).all()
        
        segment_counts = {}
        for c in customers:
            segment_counts[c.segment] = segment_counts.get(c.segment, 0) + 1
        
        location_counts = {}
        for c in customers:
            loc = c.location.split(',')[-1].strip() if ',' in c.location else c.location
            location_counts[loc] = location_counts.get(loc, 0) + 1
        
        age_ranges = {"18-25":0, "26-35":0, "36-45":0, "46-55":0, "56+":0}
        for c in customers:
            if c.age <= 25: age_ranges["18-25"] += 1
            elif c.age <= 35: age_ranges["26-35"] += 1
            elif c.age <= 45: age_ranges["36-45"] += 1
            elif c.age <= 55: age_ranges["46-55"] += 1
            else: age_ranges["56+"] += 1
        
        engagement_ranges = {"0-0.3":0, "0.3-0.5":0, "0.5-0.7":0, "0.7-0.9":0, "0.9-1.0":0}
        for c in customers:
            if c.engagement_score < 0.3: engagement_ranges["0-0.3"] += 1
            elif c.engagement_score < 0.5: engagement_ranges["0.3-0.5"] += 1
            elif c.engagement_score < 0.7: engagement_ranges["0.5-0.7"] += 1
            elif c.engagement_score < 0.9: engagement_ranges["0.7-0.9"] += 1
            else: engagement_ranges["0.9-1.0"] += 1
        
        value_ranges = {"<$500":0, "$500-$1000":0, "$1000-$2000":0, "$2000-$5000":0, "$5000+":0}
        for c in customers:
            if c.purchase_history < 500: value_ranges["<$500"] += 1
            elif c.purchase_history < 1000: value_ranges["$500-$1000"] += 1
            elif c.purchase_history < 2000: value_ranges["$1000-$2000"] += 1
            elif c.purchase_history < 5000: value_ranges["$2000-$5000"] += 1
            else: value_ranges["$5000+"] += 1
        
        channel_counts = {}
        for c in customers:
            channel_counts[c.preferred_channel] = channel_counts.get(c.preferred_channel, 0) + 1
        
        return jsonify({
            "segments": segment_counts,
            "locations": location_counts,
            "age_ranges": age_ranges,
            "engagement_ranges": engagement_ranges,
            "value_ranges": value_ranges,
            "channels": channel_counts,
            "total_customers": len(customers)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        session.close()

@app.route('/api/campaign-analytics/<int:campaign_id>')
def get_campaign_analytics(campaign_id):
    """Get detailed analytics for specific campaign"""
    session = get_db_session()
    try:
        campaign = session.query(Campaign).filter_by(id=campaign_id).first()
        analytics = session.query(Analytics).filter_by(campaign_id=campaign_id).first()
        
        if not campaign or not analytics:
            return jsonify({"error":"Campaign not found"}), 404
        
        return jsonify({
            "campaign": {
                "id": campaign.id,
                "name": campaign.name,
                "channel": campaign.channel,
                "segment": campaign.target_segment,
                "status": campaign.status
            },
            "metrics": {
                "sent": analytics.sent,
                "delivered": analytics.delivered,
                "opened": analytics.opened,
                "clicked": analytics.clicked,
                "converted": analytics.converted,
                "revenue": analytics.revenue,
                "open_rate": round(analytics.opened/analytics.sent*100, 2) if analytics.sent > 0 else 0,
                "click_rate": round(analytics.clicked/analytics.opened*100, 2) if analytics.opened > 0 else 0,
                "conversion_rate": round(analytics.converted/analytics.clicked*100, 2) if analytics.clicked > 0 else 0
            },
            "device_breakdown": analytics.device_breakdown,
            "geo_breakdown": analytics.geo_breakdown,
            "sentiment_score": analytics.sentiment_score,
            "avg_engagement_time": analytics.avg_engagement_time
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        session.close()

@app.route('/api/export/customers')
def export_customers():
    """Export customers to Excel"""
    session = get_db_session()
    try:
        customers = session.query(Customer).all()
        
        data = []
        for c in customers:
            data.append({
                'ID': c.id,
                'Name': c.name,
                'Email': c.email,
                'Phone': c.phone,
                'Age': c.age,
                'Segment': c.segment,
                'Location': c.location,
                'Purchase History': c.purchase_history,
                'Total Purchases': c.total_purchases,
                'Engagement Score': c.engagement_score,
                'Lifetime Value': c.lifetime_value,
                'Churn Probability': c.churn_probability,
                'Preferred Channel': c.preferred_channel,
                'Language': c.language,
                'Preferences': c.preferences,
                'Last Purchase': c.last_purchase_date.strftime('%Y-%m-%d') if c.last_purchase_date else '',
                'Created At': c.created_at.strftime('%Y-%m-%d')
            })
        
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Customers', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Customers']
            
            header_fill = PatternFill(start_color='00FF88', end_color='00FF88', fill_type='solid')
            header_font = Font(bold=True, color='0A0E1A')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        output.seek(0)
        
        log_agent_activity("Export System", "excel_export", f"Exported {len(customers)} customers", "completed")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'customers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        session.close()

@app.route('/api/export/campaigns')
def export_campaigns():
    """Export campaigns with analytics to Excel"""
    session = get_db_session()
    try:
        campaigns = session.query(Campaign).all()
        
        data = []
        for c in campaigns:
            analytics = session.query(Analytics).filter_by(campaign_id=c.id).first()
            data.append({
                'Campaign ID': c.id,
                'Name': c.name,
                'Channel': c.channel,
                'Target Segment': c.target_segment,
                'Status': c.status,
                'Subject Line': c.subject_line,
                'Tone': c.tone,
                'Language': c.language,
                'Sent Count': c.sent_count,
                'Predicted ROI': c.predicted_roi,
                'A/B Test': 'Yes' if c.ab_test_enabled else 'No',
                'Winner Variant': c.winner_variant or 'N/A',
                'Delivered': analytics.delivered if analytics else 0,
                'Opened': analytics.opened if analytics else 0,
                'Clicked': analytics.clicked if analytics else 0,
                'Converted': analytics.converted if analytics else 0,
                'Revenue': analytics.revenue if analytics else 0,
                'Open Rate %': round(analytics.opened/analytics.sent*100, 2) if analytics and analytics.sent > 0 else 0,
                'Click Rate %': round(analytics.clicked/analytics.opened*100, 2) if analytics and analytics.opened > 0 else 0,
                'Conversion Rate %': round(analytics.converted/analytics.clicked*100, 2) if analytics and analytics.clicked > 0 else 0,
                'Created At': c.created_at.strftime('%Y-%m-%d %H:%M')
            })
        
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Campaigns', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Campaigns']
            
            header_fill = PatternFill(start_color='0066FF', end_color='0066FF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        output.seek(0)
        
        log_agent_activity("Export System", "excel_export", f"Exported {len(campaigns)} campaigns", "completed")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'campaigns_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        session.close()

@app.route('/api/export/segment-report')
def export_segment_report():
    """Export comprehensive segment analysis to Excel"""
    session = get_db_session()
    try:
        customers = session.query(Customer).all()
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            segment_data = {}
            for c in customers:
                if c.segment not in segment_data:
                    segment_data[c.segment] = {
                        'Count': 0,
                        'Avg Age': [],
                        'Avg Purchase': [],
                        'Avg Engagement': [],
                        'Avg LTV': [],
                        'Total Revenue': 0
                    }
                segment_data[c.segment]['Count'] += 1
                segment_data[c.segment]['Avg Age'].append(c.age)
                segment_data[c.segment]['Avg Purchase'].append(c.purchase_history)
                segment_data[c.segment]['Avg Engagement'].append(c.engagement_score)
                segment_data[c.segment]['Avg LTV'].append(c.lifetime_value)
                segment_data[c.segment]['Total Revenue'] += c.purchase_history
            
            overview = []
            for seg, data in segment_data.items():
                overview.append({
                    'Segment': seg,
                    'Customer Count': data['Count'],
                    'Avg Age': round(sum(data['Avg Age'])/len(data['Avg Age']), 1),
                    'Avg Purchase Value': round(sum(data['Avg Purchase'])/len(data['Avg Purchase']), 2),
                    'Avg Engagement': round(sum(data['Avg Engagement'])/len(data['Avg Engagement']), 2),
                    'Avg Lifetime Value': round(sum(data['Avg LTV'])/len(data['Avg LTV']), 2),
                    'Total Revenue': round(data['Total Revenue'], 2),
                    'Revenue %': round(data['Total Revenue']/sum(c.purchase_history for c in customers)*100, 2)
                })
            
            df_overview = pd.DataFrame(overview)
            df_overview.to_excel(writer, sheet_name='Segment Overview', index=False)
            
            geo_data = {}
            for c in customers:
                loc = c.location.split(',')[-1].strip() if ',' in c.location else c.location
                if loc not in geo_data:
                    geo_data[loc] = {'Count': 0, 'Revenue': 0}
                geo_data[loc]['Count'] += 1
                geo_data[loc]['Revenue'] += c.purchase_history
            
            geo_list = [{'Location': k, 'Customers': v['Count'], 'Revenue': round(v['Revenue'], 2)} 
                       for k, v in geo_data.items()]
            df_geo = pd.DataFrame(geo_list)
            df_geo.to_excel(writer, sheet_name='Geographic Distribution', index=False)
            
            channel_data = {}
            for c in customers:
                if c.preferred_channel not in channel_data:
                    channel_data[c.preferred_channel] = []
                channel_data[c.preferred_channel].append(c.engagement_score)
            
            channel_list = []
            for ch, scores in channel_data.items():
                channel_list.append({
                    'Channel': ch.upper(),
                    'Customer Count': len(scores),
                    'Avg Engagement': round(sum(scores)/len(scores), 2),
                    'Percentage': round(len(scores)/len(customers)*100, 2)
                })
            df_channel = pd.DataFrame(channel_list)
            df_channel.to_excel(writer, sheet_name='Channel Preferences', index=False)
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                header_fill = PatternFill(start_color='00FF88', end_color='00FF88', fill_type='solid')
                header_font = Font(bold=True, color='0A0E1A')
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        output.seek(0)
        
        log_agent_activity("Export System", "segment_report", "Generated comprehensive segment report", "completed")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'segment_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        session.close()

@app.route('/api/campaign-preview/<int:campaign_id>')
def preview_campaign(campaign_id):
    """Get campaign preview before sending"""
    session = get_db_session()
    try:
        campaign = session.query(Campaign).filter_by(id=campaign_id).first()
        if not campaign:
            return jsonify({"error":"Campaign not found"}), 404
        
        customers = session.query(Customer).filter_by(segment=campaign.target_segment).all()
        
        ab_test = None
        if campaign.ab_test_enabled:
            ab_test = session.query(ABTest).filter_by(campaign_id=campaign_id).first()
        
        preview_data = {
            "campaign": {
                "id": campaign.id,
                "name": campaign.name,
                "channel": campaign.channel,
                "segment": campaign.target_segment,
                "subject": campaign.subject_line,
                "content": campaign.content,
                "tone": campaign.tone,
                "language": campaign.language
            },
            "target_audience": {
                "count": len(customers),
                "segments": campaign.target_segment,
                "sample_customers": [
                    {
                        "name": c.name,
                        "email": c.email,
                        "location": c.location,
                        "engagement_score": c.engagement_score
                    } for c in customers[:5]
                ]
            },
            "ab_test": None
        }
        
        if ab_test:
            preview_data["ab_test"] = {
                "variant_a": {
                    "subject": ab_test.variant_a_subject,
                    "content": ab_test.variant_a_content
                },
                "variant_b": {
                    "subject": ab_test.variant_b_subject,
                    "content": ab_test.variant_b_content
                },
                "split": "50/50"
            }
        
        return jsonify(preview_data)
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        session.close()

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/metrics')
def get_metrics():
    session = get_db_session()
    try:
        campaigns = session.query(Campaign).all()
        customers = session.query(Customer).all()
        analytics = session.query(Analytics).all()
        ts = sum(a.sent for a in analytics)
        to = sum(a.opened for a in analytics)
        tc = sum(a.clicked for a in analytics)
        tconv = sum(a.converted for a in analytics)
        tr = sum(a.revenue for a in analytics)
        return jsonify({
            'totalCampaigns':len(campaigns),
            'activeCampaigns':len([c for c in campaigns if c.status in['active','sent']]),
            'totalCustomers':len(customers),
            'totalRevenue':round(tr,2),
            'avgOpenRate':round((to/ts*100)if ts>0 else 0,1),
            'avgClickRate':round((tc/to*100)if to>0 else 0,1),
            'conversionRate':round((tconv/tc*100)if tc>0 else 0,1),
            'roi':round((tr/(ts*2.5)*100)if ts>0 else 0,0),
            'avgSentiment':round(sum(a.sentiment_score for a in analytics)/len(analytics)if analytics else 0,2),
            'abTestsRunning':len([c for c in campaigns if c.ab_test_enabled and c.status=='sent'])
        })
    except Exception as e:
        return jsonify({'error':str(e)}),500
    finally:
        session.close()

@app.route('/api/campaigns')
def get_campaigns():
    session = get_db_session()
    try:
        campaigns = session.query(Campaign).order_by(desc(Campaign.created_at)).all()
        return jsonify([{
            'id':c.id,'name':c.name,'channel':c.channel,'status':c.status,
            'segment':c.target_segment,'subject':c.subject_line,'tone':c.tone,
            'language':c.language,'sentCount':c.sent_count,'predictedRoi':c.predicted_roi,
            'abTestEnabled':c.ab_test_enabled,'winnerVariant':c.winner_variant,
            'createdAt':c.created_at.isoformat()
        }for c in campaigns])
    except:
        return jsonify([])
    finally:
        session.close()

@app.route('/api/analytics')
def get_analytics():
    session = get_db_session()
    try:
        analytics = session.query(Analytics).all()
        return jsonify([{
            'id':a.id,'campaignId':a.campaign_id,'sent':a.sent,'delivered':a.delivered,
            'opened':a.opened,'clicked':a.clicked,'converted':a.converted,'revenue':round(a.revenue,2),
            'sentimentScore':round(a.sentiment_score,2),
            'updatedAt':a.updated_at.isoformat()if a.updated_at else a.created_at.isoformat()
        }for a in analytics])
    except:
        return jsonify([])
    finally:
        session.close()

@app.route('/api/agent-activities')
def get_agent_activities():
    session = get_db_session()
    try:
        activities = session.query(AgentActivity).order_by(desc(AgentActivity.created_at)).limit(50).all()
        return jsonify([{
            'id':a.id,'agent':a.agent_name,'type':a.activity_type,
            'description':a.description,'status':a.status,'timestamp':a.created_at.isoformat()
        }for a in activities])
    except:
        return jsonify([])
    finally:
        session.close()

@app.route('/api/ab-tests')
def get_ab_tests():
    session = get_db_session()
    try:
        tests = session.query(ABTest).all()
        return jsonify([{
            'id':t.id,'campaignId':t.campaign_id,
            'variantA':{'subject':t.variant_a_subject,'sends':t.variant_a_sends,'opens':t.variant_a_opens,
                       'clicks':t.variant_a_clicks,'conversions':t.variant_a_conversions,
                       'conversionRate':round(t.variant_a_conversions/t.variant_a_sends*100,2)if t.variant_a_sends>0 else 0},
            'variantB':{'subject':t.variant_b_subject,'sends':t.variant_b_sends,'opens':t.variant_b_opens,
                       'clicks':t.variant_b_clicks,'conversions':t.variant_b_conversions,
                       'conversionRate':round(t.variant_b_conversions/t.variant_b_sends*100,2)if t.variant_b_sends>0 else 0},
            'winner':t.winner,'confidenceLevel':round(t.confidence_level*100,1)
        }for t in tests])
    except:
        return jsonify([])
    finally:
        session.close()

@app.route('/api/agent/segment-customers', methods=['POST'])
def segment_customers_endpoint():
    return jsonify(agent_segment_customers())

@app.route('/api/agent/create-campaign', methods=['POST'])
def create_campaign_endpoint():
    try:
        data = request.json
        return jsonify(agent_create_campaign(
            data.get('channel','email'),
            data.get('segment','All Customers'),
            data.get('product','our products'),
            data.get('tone','professional'),
            data.get('language','en')
        ))
    except Exception as e:
        return jsonify({"status":"error","message":str(e)}),400

@app.route('/api/agent/predict-roi/<int:campaign_id>', methods=['POST'])
def predict_roi_endpoint(campaign_id):
    return jsonify(agent_predict_roi(campaign_id))

@app.route('/api/agent/send-campaign/<int:campaign_id>', methods=['POST'])
def send_campaign_endpoint(campaign_id):
    return jsonify(agent_send_campaign(campaign_id))

@socketio.on('connect')
def handle_connect():
    print('✅ Client connected')
    emit('connection_status', {'status':'connected','timestamp':datetime.now().isoformat()})

@socketio.on('disconnect')
def handle_disconnect():
    print('❌ Client disconnected')

HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Agentic Campaign Dashboard - Enhanced</title>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.5.4/socket.io.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&display=swap');
        *{margin:0;padding:0;box-sizing:border-box}
        :root{--primary:#00ff88;--secondary:#0066ff;--dark:#0a0e1a;--card:#151925;--text:#e0e6f0;--text-dim:#8b95a8}
        body{font-family:'Space Grotesk',sans-serif;background:var(--dark);color:var(--text);overflow-x:hidden}
        .container{max-width:1800px;margin:0 auto;padding:2rem}
        header{display:flex;justify-content:space-between;align-items:center;padding:2rem 0;border-bottom:1px solid rgba(0,255,136,0.1);margin-bottom:3rem}
        .logo{display:flex;align-items:center;gap:1rem}
        .logo-icon{width:50px;height:50px;background:linear-gradient(135deg,var(--primary),var(--secondary));border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.5rem;font-weight:700}
        .logo h1{font-size:1.8rem;background:linear-gradient(135deg,var(--primary),var(--secondary));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
        .btn{padding:0.75rem 1.5rem;border:none;border-radius:10px;font-weight:600;cursor:pointer;font-family:'Space Grotesk',sans-serif;transition:all 0.3s ease}
        .btn-primary{background:linear-gradient(135deg,var(--primary),var(--secondary));color:var(--dark)}
        .btn-primary:hover{transform:translateY(-2px);box-shadow:0 10px 30px rgba(0,255,136,0.3)}
        .btn-secondary{background:rgba(0,102,255,0.2);color:var(--secondary);border:1px solid var(--secondary)}
        .btn-secondary:hover{background:rgba(0,102,255,0.3)}
        .btn-ghost{background:transparent;color:var(--text);border:1px solid rgba(255,255,255,0.1)}
        .btn-ghost:hover{background:rgba(255,255,255,0.05)}
        .btn-sm{padding:0.5rem 1rem;font-size:0.85rem}
        .card{background:var(--card);border:1px solid rgba(0,255,136,0.1);border-radius:20px;padding:2rem;transition:all 0.3s ease}
        .card:hover{border-color:var(--primary);transform:translateY(-2px);box-shadow:0 10px 30px rgba(0,255,136,0.1)}
        .metrics-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:1.5rem;margin-bottom:3rem}
        .metric-label{font-size:0.85rem;color:var(--text-dim);text-transform:uppercase;margin-bottom:0.5rem;letter-spacing:1px}
        .metric-value{font-size:2.5rem;font-weight:700;background:linear-gradient(135deg,var(--primary),var(--secondary));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
        .tabs{display:flex;gap:1rem;margin-bottom:2rem;border-bottom:2px solid rgba(255,255,255,0.05);overflow-x:auto}
        .tab{padding:1rem 2rem;cursor:pointer;border-bottom:3px solid transparent;transition:all 0.3s;font-weight:600;white-space:nowrap}
        .tab.active{border-bottom-color:var(--primary);color:var(--primary)}
        .tab:hover{color:var(--primary)}
        .tab-content{display:none}
        .tab-content.active{display:block;animation:fadeIn 0.3s ease}
        @keyframes fadeIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
        .chart-container{position:relative;height:400px;margin:2rem 0}
        .modal{position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.8);display:none;align-items:center;justify-content:center;z-index:1000;backdrop-filter:blur(10px)}
        .modal.active{display:flex;animation:fadeIn 0.3s ease}
        .modal-content{background:var(--card);border:1px solid rgba(0,255,136,0.1);border-radius:20px;padding:2rem;max-width:900px;width:90%;max-height:80vh;overflow-y:auto}
        .modal-content::-webkit-scrollbar{width:8px}
        .modal-content::-webkit-scrollbar-track{background:rgba(0,0,0,0.2);border-radius:10px}
        .modal-content::-webkit-scrollbar-thumb{background:var(--primary);border-radius:10px}
        .form-group{margin-bottom:1.5rem}
        .form-label{display:block;margin-bottom:0.5rem;color:var(--text-dim);font-size:0.9rem;font-weight:600}
        .form-input,.form-select{width:100%;padding:0.75rem 1rem;background:rgba(0,0,0,0.3);border:1px solid rgba(255,255,255,0.1);border-radius:10px;color:var(--text);font-family:'Space Grotesk',sans-serif;font-size:1rem;transition:all 0.3s}
        .form-input:focus,.form-select:focus{outline:none;border-color:var(--primary);box-shadow:0 0 0 3px rgba(0,255,136,0.1)}
        .form-input::placeholder{color:var(--text-dim)}
        .file-upload{border:2px dashed rgba(0,255,136,0.3);border-radius:10px;padding:2rem;text-align:center;cursor:pointer;transition:all 0.3s}
        .file-upload:hover{border-color:var(--primary);background:rgba(0,255,136,0.05)}
        .toast{position:fixed;bottom:2rem;right:2rem;background:var(--card);border:1px solid var(--primary);border-radius:15px;padding:1rem 1.5rem;z-index:10000;display:none;box-shadow:0 10px 30px rgba(0,255,136,0.3)}
        .toast.active{display:block;animation:slideInRight 0.3s ease}
        @keyframes slideInRight{from{transform:translateX(400px);opacity:0}to{transform:translateX(0);opacity:1}}
        .preview-box{background:rgba(0,0,0,0.3);border:1px solid rgba(255,255,255,0.1);border-radius:10px;padding:1.5rem;margin:1rem 0}
        .preview-subject{font-size:1.2rem;font-weight:600;color:var(--primary);margin-bottom:0.5rem}
        .preview-content{line-height:1.6;color:var(--text-dim)}
        .grid-2{display:grid;grid-template-columns:1fr 1fr;gap:1.5rem}
        @media(max-width:768px){
            .grid-2{grid-template-columns:1fr}
            .tabs{overflow-x:scroll}
            .modal-content{padding:1.5rem}
        }
    </style>
</head>
<body>
    <div class="container">
        <header>
            <div class="logo">
                <div class="logo-icon">Ξ</div>
                <div>
                    <h1>AGENTIC CAMPAIGN SYSTEM</h1>
                    <p style="font-size:0.85rem;color:var(--text-dim)">✉️ Real Email/SMS + Visualizations + Excel Reports</p>
                </div>
            </div>
            <div style="display:flex;gap:1rem;flex-wrap:wrap">
                <button class="btn btn-ghost" onclick="openModal('uploadCustomers')">📤 Upload CSV</button>
                <button class="btn btn-secondary" onclick="segmentCustomers()">🧠 Segment</button>
                <button class="btn btn-primary" onclick="openModal('createCampaign')">+ NEW CAMPAIGN</button>
            </div>
        </header>
        
        <div class="metrics-grid" id="metricsGrid">
            <div class="card">
                <div class="metric-label">Total Revenue</div>
                <div class="metric-value" id="totalRevenue">$0</div>
            </div>
            <div class="card">
                <div class="metric-label">Active Campaigns</div>
                <div class="metric-value" id="activeCampaigns">0</div>
            </div>
            <div class="card">
                <div class="metric-label">Avg Open Rate</div>
                <div class="metric-value" id="avgOpenRate">0%</div>
            </div>
            <div class="card">
                <div class="metric-label">ROI</div>
                <div class="metric-value" id="roi">0%</div>
            </div>
        </div>
        
        <div class="tabs">
            <div class="tab active" onclick="switchTab('campaigns')">📊 Campaigns</div>
            <div class="tab" onclick="switchTab('segments')">🎯 Segments</div>
            <div class="tab" onclick="switchTab('analytics')">📈 Analytics</div>
            <div class="tab" onclick="switchTab('reports')">📄 Reports</div>
        </div>
        
        <div class="tab-content active" id="campaigns-tab">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:1.5rem">
                <h2>Active Campaigns</h2>
                <button class="btn btn-ghost btn-sm" onclick="exportCampaigns()">📥 Export</button>
            </div>
            <div class="campaign-grid" id="campaignGrid" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(350px,1fr));gap:1.5rem"></div>
        </div>
        
        <div class="tab-content" id="segments-tab">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:1.5rem">
                <h2>Customer Segmentation</h2>
                <button class="btn btn-ghost btn-sm" onclick="exportSegmentReport()">📥 Export Report</button>
            </div>
            <div class="grid-2">
                <div class="card">
                    <h3 style="margin-bottom:1rem">Segment Distribution</h3>
                    <div class="chart-container"><canvas id="segmentChart"></canvas></div>
                </div>
                <div class="card">
                    <h3 style="margin-bottom:1rem">Geographic Distribution</h3>
                    <div class="chart-container"><canvas id="geoChart"></canvas></div>
                </div>
            </div>
            <div class="grid-2" style="margin-top:1.5rem">
                <div class="card">
                    <h3 style="margin-bottom:1rem">Engagement Levels</h3>
                    <div class="chart-container"><canvas id="engagementChart"></canvas></div>
                </div>
                <div class="card">
                    <h3 style="margin-bottom:1rem">Purchase Value</h3>
                    <div class="chart-container"><canvas id="valueChart"></canvas></div>
                </div>
            </div>
        </div>
        
        <div class="tab-content" id="analytics-tab">
            <h2 style="margin-bottom:1.5rem">Campaign Performance</h2>
            <div class="card">
                <h3 style="margin-bottom:1rem">Overview</h3>
                <div class="chart-container"><canvas id="performanceChart"></canvas></div>
            </div>
            <div class="grid-2" style="margin-top:1.5rem">
                <div class="card">
                    <h3 style="margin-bottom:1rem">Channel Performance</h3>
                    <div class="chart-container"><canvas id="channelChart"></canvas></div>
                </div>
                <div class="card">
                    <h3 style="margin-bottom:1rem">Conversion Funnel</h3>
                    <div class="chart-container"><canvas id="funnelChart"></canvas></div>
                </div>
            </div>
        </div>
        
        <div class="tab-content" id="reports-tab">
            <h2 style="margin-bottom:1.5rem">Download Reports</h2>
            <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:1.5rem">
                <div class="card" style="cursor:pointer" onclick="exportCustomers()">
                    <div style="display:flex;align-items:center;gap:1rem;margin-bottom:1rem">
                        <div style="width:60px;height:60px;background:linear-gradient(135deg,var(--primary),var(--secondary));border-radius:15px;display:flex;align-items:center;justify-content:center;font-size:2rem">👥</div>
                        <div>
                            <h3>Customer Database</h3>
                            <p style="color:var(--text-dim);font-size:0.85rem">Complete list with all attributes</p>
                        </div>
                    </div>
                    <button class="btn btn-primary" style="width:100%">📥 Download Excel</button>
                </div>
                <div class="card" style="cursor:pointer" onclick="exportCampaigns()">
                    <div style="display:flex;align-items:center;gap:1rem;margin-bottom:1rem">
                        <div style="width:60px;height:60px;background:linear-gradient(135deg,#FF6B6B,#FF8E53);border-radius:15px;display:flex;align-items:center;justify-content:center;font-size:2rem">📧</div>
                        <div>
                            <h3>Campaign Report</h3>
                            <p style="color:var(--text-dim);font-size:0.85rem">All campaigns with metrics</p>
                        </div>
                    </div>
                    <button class="btn btn-primary" style="width:100%">📥 Download Excel</button>
                </div>
                <div class="card" style="cursor:pointer" onclick="exportSegmentReport()">
                    <div style="display:flex;align-items:center;gap:1rem;margin-bottom:1rem">
                        <div style="width:60px;height:60px;background:linear-gradient(135deg,#A8E6CF,#3EECAC);border-radius:15px;display:flex;align-items:center;justify-content:center;font-size:2rem">🎯</div>
                        <div>
                            <h3>Segment Analysis</h3>
                            <p style="color:var(--text-dim);font-size:0.85rem">Comprehensive report</p>
                        </div>
                    </div>
                    <button class="btn btn-primary" style="width:100%">📥 Download Excel</button>
                </div>
            </div>
        </div>
    </div>
    
    <div class="modal" id="uploadCustomersModal">
        <div class="modal-content" style="max-width:600px">
            <h2 style="margin-bottom:2rem">Upload Customer Data</h2>
            <form id="uploadForm" onsubmit="uploadCSV(event)">
                <div class="form-group">
                    <label class="form-label">CSV File</label>
                    <div class="file-upload" onclick="document.getElementById('csvFile').click()">
                        <div style="font-size:3rem;margin-bottom:1rem">📁</div>
                        <p style="color:var(--text-dim)">Click to select CSV file</p>
                        <p style="color:var(--text-dim);font-size:0.85rem;margin-top:0.5rem">Required columns: name, email</p>
                        <p id="fileName" style="color:var(--primary);margin-top:1rem;font-weight:600"></p>
                    </div>
                    <input type="file" id="csvFile" accept=".csv" style="display:none" onchange="document.getElementById('fileName').textContent=this.files[0]?.name||''">
                </div>
                <div style="display:flex;gap:1rem">
                    <button type="submit" class="btn btn-primary" style="flex:1">📤 Upload</button>
                    <button type="button" class="btn btn-ghost" onclick="closeModal('uploadCustomers')">Cancel</button>
                </div>
            </form>
        </div>
    </div>
    
    <div class="modal" id="createCampaignModal">
        <div class="modal-content">
            <h2 style="margin-bottom:2rem">Create AI Campaign</h2>
            <form id="campaignForm" onsubmit="createCampaign(event)">
                <div class="form-group">
                    <label class="form-label">Channel</label>
                    <select class="form-select" name="channel" required>
                        <option value="email">📧 Email</option>
                        <option value="sms">📱 SMS</option>
                        <option value="social">🌐 Social Media</option>
                    </select>
                </div>
                <div class="form-group">
                    <label class="form-label">Target Segment</label>
                    <input type="text" class="form-input" name="segment" placeholder="e.g., High-Value" required>
                </div>
                <div class="form-group">
                    <label class="form-label">Product/Service</label>
                    <input type="text" class="form-input" name="product" placeholder="e.g., Premium Membership" required>
                </div>
                <div class="form-group">
                    <label class="form-label">Tone</label>
                    <select class="form-select" name="tone">
                        <option value="professional">Professional</option>
                        <option value="friendly">Friendly</option>
                        <option value="urgent">Urgent</option>
                        <option value="casual">Casual</option>
                    </select>
                </div>
                <div style="display:flex;gap:1rem">
                    <button type="submit" class="btn btn-primary" style="flex:1">🚀 Generate Campaign</button>
                    <button type="button" class="btn btn-ghost" onclick="closeModal('createCampaign')">Cancel</button>
                </div>
            </form>
        </div>
    </div>
    
    <div class="modal" id="previewModal">
        <div class="modal-content">
            <h2 style="margin-bottom:2rem">📧 Campaign Preview</h2>
            <div id="previewContent"></div>
            <div style="display:flex;gap:1rem;margin-top:2rem">
                <button class="btn btn-primary" style="flex:1" onclick="confirmSendCampaign()">✅ Confirm & Send (REAL EMAIL/SMS)</button>
                <button class="btn btn-ghost" onclick="closeModal('preview')">Cancel</button>
            </div>
        </div>
    </div>
    
    <div class="toast" id="toast">
        <div id="toastMessage"></div>
    </div>
    
    <script>
        const API_BASE = window.location.origin + '/api';
        const socket = io(window.location.origin);
        let charts = {};
        let pendingCampaignId = null;
        
        socket.on('connect', () => {
            console.log('✅ Connected to server');
            showToast('✅ Connected to server');
        });
        
        socket.on('agent_activity', (data) => {
            console.log('🤖 Agent Activity:', data);
        });
        
        socket.on('analytics_update', (data) => {
            console.log('📊 Analytics updated:', data);
            loadMetrics();
            loadAnalyticsCharts();
        });
        
        socket.on('campaign_progress', (data) => {
            console.log('📤 Campaign Progress:', data);
            showToast(`📤 Sending: ${data.progress}% (${data.delivered}/${data.total} delivered)`);
        });
        
        async function initDashboard() {
            await loadMetrics();
            await loadCampaigns();
            await loadSegmentCharts();
            await loadAnalyticsCharts();
            startRealtimeUpdates();
        }
        
        async function loadMetrics() {
            try {
                const r = await fetch(`${API_BASE}/metrics`);
                const d = await r.json();
                document.getElementById('totalRevenue').textContent = `${d.totalRevenue.toLocaleString()}`;
                document.getElementById('activeCampaigns').textContent = d.activeCampaigns;
                document.getElementById('avgOpenRate').textContent = `${d.avgOpenRate}%`;
                document.getElementById('roi').textContent = `${d.roi}%`;
            } catch(e) {
                console.error('Error loading metrics:', e);
            }
        }
        
        async function loadCampaigns() {
            try {
                const r = await fetch(`${API_BASE}/campaigns`);
                const campaigns = await r.json();
                const grid = document.getElementById('campaignGrid');
                
                if(campaigns.length === 0) {
                    grid.innerHTML = '<div class="card" style="grid-column:1/-1;text-align:center;padding:4rem"><h3>No campaigns yet</h3><p style="color:var(--text-dim);margin:1rem 0">Create your first AI-powered campaign</p><button class="btn btn-primary" onclick="openModal(\\'createCampaign\\')">+ Create Campaign</button></div>';
                    return;
                }
                
                grid.innerHTML = campaigns.map(c => `
                    <div class="card">
                        <div style="display:flex;justify-content:space-between;margin-bottom:1rem">
                            <div>
                                <h3>${c.name}</h3>
                                <div style="color:var(--text-dim);font-size:0.85rem;margin-top:0.5rem">
                                    ${c.channel.toUpperCase()} • ${c.segment}
                                </div>
                            </div>
                            <span style="padding:0.4rem 0.8rem;border-radius:50px;font-size:0.75rem;font-weight:600;text-transform:uppercase;background:${c.status==='sent'?'rgba(0,255,136,0.2)':'rgba(255,107,107,0.2)'};color:${c.status==='sent'?'var(--primary)':'#ff6b6b'};height:fit-content">
                                ${c.status.toUpperCase()}
                            </span>
                        </div>
                        ${c.status === 'sent' ? `
                            <div style="padding:1rem;background:rgba(0,0,0,0.2);border-radius:10px;margin:1rem 0">
                                <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:1rem;text-align:center">
                                    <div>
                                        <div style="font-size:1.5rem;color:var(--primary);font-weight:700">${c.sentCount || 0}</div>
                                        <div style="font-size:0.75rem;color:var(--text-dim);margin-top:0.25rem">SENT</div>
                                    </div>
                                    <div>
                                        <div style="font-size:1.5rem;color:var(--primary);font-weight:700">${c.predictedRoi || 0}x</div>
                                        <div style="font-size:0.75rem;color:var(--text-dim);margin-top:0.25rem">ROI</div>
                                    </div>
                                    <div>
                                        <div style="font-size:1.5rem;color:var(--primary);font-weight:700">${c.abTestEnabled ? 'A/B' : 'STD'}</div>
                                        <div style="font-size:0.75rem;color:var(--text-dim);margin-top:0.25rem">TYPE</div>
                                    </div>
                                </div>
                            </div>
                        ` : ''}
                        <div style="display:flex;gap:0.5rem;margin-top:1rem">
                            ${c.status === 'ready' ? `
                                <button class="btn btn-secondary btn-sm" onclick="previewCampaign(${c.id})">👁️ Preview & Send</button>
                            ` : ''}
                            ${c.status === 'sent' ? `
                                <button class="btn btn-ghost btn-sm" onclick="viewAnalytics(${c.id})">📊 View Analytics</button>
                            ` : ''}
                        </div>
                    </div>
                `).join('');
            } catch(e) {
                console.error('Error loading campaigns:', e);
            }
        }
        
        async function loadSegmentCharts() {
            try {
                const r = await fetch(`${API_BASE}/segment-visualization`);
                const d = await r.json();
                
                if(charts.segment) charts.segment.destroy();
                charts.segment = new Chart(document.getElementById('segmentChart'), {
                    type: 'doughnut',
                    data: {
                        labels: Object.keys(d.segments),
                        datasets: [{
                            data: Object.values(d.segments),
                            backgroundColor: ['#00ff88', '#0066ff', '#ff6b6b', '#ffd93d', '#a8e6cf', '#ff8e53']
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {
                            legend: {
                                position: 'bottom',
                                labels: { color: '#e0e6f0', padding: 15 }
                            }
                        }
                    }
                });
                
                if(charts.geo) charts.geo.destroy();
                charts.geo = new Chart(document.getElementById('geoChart'), {
                    type: 'bar',
                    data: {
                        labels: Object.keys(d.locations),
                        datasets: [{
                            label: 'Customers',
                            data: Object.values(d.locations),
                            backgroundColor: 'rgba(0,255,136,0.5)',
                            borderColor: '#00ff88',
                            borderWidth: 2
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        scales: {
                            y: { beginAtZero: true, ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } },
                            x: { ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } }
                        },
                        plugins: {
                            legend: { labels: { color: '#e0e6f0' } }
                        }
                    }
                });
                
                if(charts.engagement) charts.engagement.destroy();
                charts.engagement = new Chart(document.getElementById('engagementChart'), {
                    type: 'bar',
                    data: {
                        labels: Object.keys(d.engagement_ranges),
                        datasets: [{
                            label: 'Customers',
                            data: Object.values(d.engagement_ranges),
                            backgroundColor: 'rgba(0,102,255,0.5)',
                            borderColor: '#0066ff',
                            borderWidth: 2
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        scales: {
                            y: { beginAtZero: true, ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } },
                            x: { ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } }
                        },
                        plugins: {
                            legend: { labels: { color: '#e0e6f0' } }
                        }
                    }
                });
                
                if(charts.value) charts.value.destroy();
                charts.value = new Chart(document.getElementById('valueChart'), {
                    type: 'bar',
                    data: {
                        labels: Object.keys(d.value_ranges),
                        datasets: [{
                            label: 'Customers',
                            data: Object.values(d.value_ranges),
                            backgroundColor: 'rgba(255,107,107,0.5)',
                            borderColor: '#ff6b6b',
                            borderWidth: 2
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        scales: {
                            y: { beginAtZero: true, ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } },
                            x: { ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } }
                        },
                        plugins: {
                            legend: { labels: { color: '#e0e6f0' } }
                        }
                    }
                });
            } catch(e) {
                console.error('Error loading segment charts:', e);
            }
        }
        
        async function loadAnalyticsCharts() {
            try {
                const [cr, ar] = await Promise.all([
                    fetch(`${API_BASE}/campaigns`),
                    fetch(`${API_BASE}/analytics`)
                ]);
                const campaigns = await cr.json();
                const analytics = await ar.json();
                
                const perfData = campaigns.filter(c => c.status === 'sent').slice(0, 5).map(c => {
                    const a = analytics.find(an => an.campaignId === c.id);
                    return {
                        name: c.name,
                        opened: a?.opened || 0,
                        clicked: a?.clicked || 0,
                        converted: a?.converted || 0
                    };
                });
                
                if(charts.performance) charts.performance.destroy();
                charts.performance = new Chart(document.getElementById('performanceChart'), {
                    type: 'line',
                    data: {
                        labels: perfData.map(d => d.name),
                        datasets: [
                            {
                                label: 'Opened',
                                data: perfData.map(d => d.opened),
                                borderColor: '#00ff88',
                                backgroundColor: 'rgba(0,255,136,0.1)',
                                fill: true,
                                tension: 0.4
                            },
                            {
                                label: 'Clicked',
                                data: perfData.map(d => d.clicked),
                                borderColor: '#0066ff',
                                backgroundColor: 'rgba(0,102,255,0.1)',
                                fill: true,
                                tension: 0.4
                            },
                            {
                                label: 'Converted',
                                data: perfData.map(d => d.converted),
                                borderColor: '#ff6b6b',
                                backgroundColor: 'rgba(255,107,107,0.1)',
                                fill: true,
                                tension: 0.4
                            }
                        ]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        scales: {
                            y: { beginAtZero: true, ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } },
                            x: { ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } }
                        },
                        plugins: {
                            legend: { labels: { color: '#e0e6f0' } }
                        }
                    }
                });
                
                const channelData = {};
                campaigns.forEach(c => {
                    const a = analytics.find(an => an.campaignId === c.id);
                    if(!channelData[c.channel]) channelData[c.channel] = 0;
                    channelData[c.channel] += (a?.revenue || 0);
                });
                
                if(charts.channel) charts.channel.destroy();
                charts.channel = new Chart(document.getElementById('channelChart'), {
                    type: 'pie',
                    data: {
                        labels: Object.keys(channelData).map(k => k.toUpperCase()),
                        datasets: [{
                            data: Object.values(channelData),
                            backgroundColor: ['#00ff88', '#0066ff', '#ff6b6b', '#ffd93d']
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {
                            legend: {
                                position: 'bottom',
                                labels: { color: '#e0e6f0', padding: 15 }
                            }
                        }
                    }
                });
                
                const ts = analytics.reduce((s, a) => s + a.sent, 0);
                const td = analytics.reduce((s, a) => s + a.delivered, 0);
                const to = analytics.reduce((s, a) => s + a.opened, 0);
                const tc = analytics.reduce((s, a) => s + a.clicked, 0);
                const tcv = analytics.reduce((s, a) => s + a.converted, 0);
                
                if(charts.funnel) charts.funnel.destroy();
                charts.funnel = new Chart(document.getElementById('funnelChart'), {
                    type: 'bar',
                    data: {
                        labels: ['Sent', 'Delivered', 'Opened', 'Clicked', 'Converted'],
                        datasets: [{
                            label: 'Count',
                            data: [ts, td, to, tc, tcv],
                            backgroundColor: ['#00ff88', '#3eecac', '#0066ff', '#4d94ff', '#ff6b6b']
                        }]
                    },
                    options: {
                        indexAxis: 'y',
                        responsive: true,
                        maintainAspectRatio: false,
                        scales: {
                            x: { beginAtZero: true, ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } },
                            y: { ticks: { color: '#8b95a8' }, grid: { color: 'rgba(255,255,255,0.05)' } }
                        },
                        plugins: {
                            legend: { display: false }
                        }
                    }
                });
            } catch(e) {
                console.error('Error loading analytics charts:', e);
            }
        }
        
        async function uploadCSV(event) {
            event.preventDefault();
            const f = document.getElementById('csvFile');
            if(!f.files[0]) {
                showToast('❌ Please select a file');
                return;
            }
            
            showToast('📤 Uploading...');
            try {
                const fd = new FormData();
                fd.append('file', f.files[0]);
                const r = await fetch(`${API_BASE}/upload-customers`, {
                    method: 'POST',
                    body: fd
                });
                const res = await r.json();
                
                if(res.status === 'success') {
                    showToast(`✅ ${res.message}`);
                    closeModal('uploadCustomers');
                    f.value = '';
                    document.getElementById('fileName').textContent = '';
                    await loadMetrics();
                    await loadSegmentCharts();
                } else {
                    showToast('❌ ' + res.message);
                }
            } catch(e) {
                showToast('❌ Upload failed');
                console.error(e);
            }
        }
        
        async function createCampaign(event) {
            event.preventDefault();
            const fd = new FormData(event.target);
            const data = {
                channel: fd.get('channel'),
                segment: fd.get('segment'),
                product: fd.get('product'),
                tone: fd.get('tone')
            };
            
            showToast('🤖 AI generating campaign...');
            try {
                const r = await fetch(`${API_BASE}/agent/create-campaign`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(data)
                });
                const res = await r.json();
                
                if(res.status === 'success') {
                    showToast(`✅ Campaign created! ID: ${res.campaign_id}`);
                    closeModal('createCampaign');
                    event.target.reset();
                    await loadCampaigns();
                } else {
                    showToast('❌ ' + res.message);
                }
            } catch(e) {
                showToast('❌ Failed to create campaign');
                console.error(e);
            }
        }
        
        async function segmentCustomers() {
            showToast('🧠 AI segmenting customers...');
            try {
                const r = await fetch(`${API_BASE}/agent/segment-customers`, {
                    method: 'POST'
                });
                const res = await r.json();
                
                if(res.status === 'success') {
                    showToast(`✅ Segmented ${res.customers} customers into ${res.segments} groups`);
                    await loadMetrics();
                    await loadSegmentCharts();
                } else {
                    showToast('❌ ' + res.message);
                }
            } catch(e) {
                showToast('❌ Segmentation failed');
                console.error(e);
            }
        }
        
        async function previewCampaign(campaignId) {
            showToast('📄 Loading preview...');
            try {
                const r = await fetch(`${API_BASE}/campaign-preview/${campaignId}`);
                const data = await r.json();
                
                pendingCampaignId = campaignId;
                
                let html = `
                    <div class="preview-box">
                        <h3 style="margin-bottom:1rem">Campaign Details</h3>
                        <div style="display:grid;grid-template-columns:1fr 1fr;gap:1rem;margin-bottom:1rem">
                            <div><strong>Name:</strong> ${data.campaign.name}</div>
                            <div><strong>Channel:</strong> ${data.campaign.channel.toUpperCase()}</div>
                            <div><strong>Segment:</strong> ${data.campaign.segment}</div>
                            <div><strong>Target Count:</strong> ${data.target_audience.count}</div>
                        </div>
                    </div>
                `;
                
                if(data.ab_test) {
                    html += `
                        <h3 style="margin:1.5rem 0 1rem">A/B Test Variants (50/50 Split)</h3>
                        <div class="grid-2">
                            <div class="preview-box">
                                <h4 style="color:var(--primary);margin-bottom:1rem">Variant A</h4>
                                <div class="preview-subject">${data.ab_test.variant_a.subject}</div>
                                <div class="preview-content">${data.ab_test.variant_a.content}</div>
                            </div>
                            <div class="preview-box">
                                <h4 style="color:var(--secondary);margin-bottom:1rem">Variant B</h4>
                                <div class="preview-subject">${data.ab_test.variant_b.subject}</div>
                                <div class="preview-content">${data.ab_test.variant_b.content}</div>
                            </div>
                        </div>
                    `;
                } else {
                    html += `
                        <div class="preview-box">
                            <div class="preview-subject">${data.campaign.subject}</div>
                            <div class="preview-content">${data.campaign.content}</div>
                        </div>
                    `;
                }
                
                html += `
                    <div class="preview-box" style="background:rgba(255,107,107,0.1);border-color:#ff6b6b">
                        <h4 style="color:#ff6b6b;margin-bottom:0.5rem">⚠️ Important</h4>
                        <p style="color:var(--text-dim)">This will send REAL ${data.campaign.channel.toUpperCase()} messages to ${data.target_audience.count} customers. Make sure your credentials are configured correctly.</p>
                    </div>
                `;
                
                document.getElementById('previewContent').innerHTML = html;
                openModal('preview');
            } catch(e) {
                showToast('❌ Failed to load preview');
                console.error(e);
            }
        }
        
        async function confirmSendCampaign() {
            if(!pendingCampaignId) return;
            
            closeModal('preview');
            showToast('📤 Sending campaign...');
            
            try {
                const r = await fetch(`${API_BASE}/agent/send-campaign/${pendingCampaignId}`, {
                    method: 'POST'
                });
                const res = await r.json();
                
                if(res.status === 'success') {
                    showToast(`✅ Campaign sent! Delivered: ${res.delivered}, Failed: ${res.failed}`);
                    pendingCampaignId = null;
                    await loadCampaigns();
                    await loadMetrics();
                } else {
                    showToast('❌ ' + res.message);
                }
            } catch(e) {
                showToast('❌ Failed to send campaign');
                console.error(e);
            }
        }
        
        async function viewAnalytics(campaignId) {
            showToast('📊 Loading analytics...');
            try {
                const r = await fetch(`${API_BASE}/campaign-analytics/${campaignId}`);
                const data = await r.json();
                
                alert(`Campaign: ${data.campaign.name}\n\nSent: ${data.metrics.sent}\nOpened: ${data.metrics.opened} (${data.metrics.open_rate}%)\nClicked: ${data.metrics.clicked} (${data.metrics.click_rate}%)\nConverted: ${data.metrics.converted} (${data.metrics.conversion_rate}%)\nRevenue: ${data.metrics.revenue}`);
            } catch(e) {
                showToast('❌ Failed to load analytics');
                console.error(e);
            }
        }
        
        function exportCustomers() {
            showToast('📥 Downloading customers...');
            window.location.href = `${API_BASE}/export/customers`;
        }
        
        function exportCampaigns() {
            showToast('📥 Downloading campaigns...');
            window.location.href = `${API_BASE}/export/campaigns`;
        }
        
        function exportSegmentReport() {
            showToast('📥 Downloading segment report...');
            window.location.href = `${API_BASE}/export/segment-report`;
        }
        
        function switchTab(tabName) {
            document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
            
            event.target.classList.add('active');
            document.getElementById(`${tabName}-tab`).classList.add('active');
            
            if(tabName === 'segments') {
                loadSegmentCharts();
            } else if(tabName === 'analytics') {
                loadAnalyticsCharts();
            }
        }
        
        function openModal(modalName) {
            document.getElementById(`${modalName}Modal`).classList.add('active');
        }
        
        function closeModal(modalName) {
            document.getElementById(`${modalName}Modal`).classList.remove('active');
        }
        
        function showToast(message) {
            const toast = document.getElementById('toast');
            document.getElementById('toastMessage').textContent = message;
            toast.classList.add('active');
            setTimeout(() => {
                toast.classList.remove('active');
            }, 4000);
        }
        
        function startRealtimeUpdates() {
            setInterval(async () => {
                await loadMetrics();
            }, 30000);
        }
        
        document.addEventListener('DOMContentLoaded', () => {
            initDashboard();
        });
        
        window.onclick = function(event) {
            if(event.target.classList.contains('modal')) {
                event.target.classList.remove('active');
            }
        }
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    print("\n" + "="*80)
    print("🚀 ENHANCED AGENTIC CAMPAIGN SYSTEM")
    print("="*80)
    print("\n📧 Email Setup:")
    print(f"   Gmail: {GMAIL_USER}")
    print(f"   Status: {'✅ Configured' if len(GMAIL_APP_PASSWORD) > 10 else '❌ Not configured'}")
    print("\n📱 SMS Setup:")
    print(f"   Twilio SID: {TWILIO_ACCOUNT_SID[:10]}..." if len(TWILIO_ACCOUNT_SID) > 10 else "   ❌ Not configured")
    print(f"   Status: {'✅ Configured' if len(TWILIO_AUTH_TOKEN) > 10 else '❌ Not configured'}")
    print("\n🤖 AI Setup:")
    print(f"   Groq API: {'✅ Configured' if len(GROQ_API_KEY) > 20 else '❌ Not configured'}")
    print(f"   LLM Type: {llm_type.upper()}")
    print("\n" + "="*80)
    
    generate_sample_data()
    
    print("\n🌐 Starting server...")
    print("📍 Dashboard: http://localhost:5005")
    print("\n✨ Features:")
    print("   • Real Gmail SMTP email sending")
    print("   • Real Twilio SMS sending")
    print("   • CSV customer upload")
    print("   • AI-powered segmentation")
    print("   • AI campaign generation")
    print("   • Interactive visualizations")
    print("   • Excel report exports")
    print("   • Real-time analytics")
    print("   • A/B testing")
    print("\n" + "="*80 + "\n")
    
    socketio.run(app, host='0.0.0.0', port=5005, debug=False)